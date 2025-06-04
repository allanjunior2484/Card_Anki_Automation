import requests
from gtts import gTTS
import base64
import hashlib
import os
from openpyxl import load_workbook

def gerar_nome_audio(texto):
    hash_nome = hashlib.md5(texto.encode()).hexdigest()
    return f"audio_{hash_nome}.mp3"

def gerar_audio(texto, lang="en", tld="com"):
    nome_arquivo = gerar_nome_audio(texto)
    if not os.path.exists(nome_arquivo):
        tts = gTTS(texto, lang=lang, tld=tld)
        tts.save(nome_arquivo)
    return nome_arquivo

def enviar_audio_para_anki(nome_arquivo):
    with open(nome_arquivo, "rb") as f:
        audio_data = f.read()
    audio_b64 = base64.b64encode(audio_data).decode('utf-8')
    payload = {
        "action": "storeMediaFile",
        "version": 6,
        "params": {
            "filename": nome_arquivo,
            "data": audio_b64
        }
    }
    return requests.post("http://localhost:8765", json=payload).json()

def criar_nota(front, back, audio_filename, deck_name):
    nota = {
        "action": "addNote",
        "version": 6,
        "params": {
            "note": {
                "deckName": deck_name,
                "modelName": "Basic",
                "fields": {
                    "Front": f"{front} [sound:{audio_filename}]",
                    "Back": back
                },
                "options": {
                    "allowDuplicate": False
                },
                "tags": ["excel"]
            }
        }
    }
    return requests.post("http://localhost:8765", json=nota).json()

def processar_planilha(arquivo_excel, deck_name):
    # Verifica se o baralho existe
    resposta = requests.post("http://localhost:8765", json={
        "action": "deckNames",
        "version": 6
    }).json()

    if deck_name not in resposta.get("result", []):
        print(f"❌ Baralho '{deck_name}' não encontrado.")
        return "baralho_nao_encontrado"

    wb = load_workbook(arquivo_excel)
    sheet = wb.active

    for linha in sheet.iter_rows(min_row=2, values_only=True):
        front, back = linha
        if not front or not back:
            continue
        print(f"Adicionando: {front} -> {back}")
        nome_audio = gerar_audio(front)
        enviar_audio_para_anki(nome_audio)
        criar_nota(front, back, nome_audio, deck_name)

    print("✅ Todos os cartões foram adicionados com sucesso!")

    sheet.delete_rows(2, sheet.max_row)
    wb.save(arquivo_excel)
    print("🧹 Planilha limpa e pronta para novos cartões.")
    return "sucesso"

# Execução direta
if __name__ == "__main__":
    if os.path.exists("cartoes.xlsx"):
        nome_baralho = input("Digite o nome do baralho do Anki: ")
        processar_planilha("cartoes.xlsx", nome_baralho)
    else:
        print("❌ Arquivo cartoes.xlsx não encontrado. Coloque na mesma pasta do script.")
